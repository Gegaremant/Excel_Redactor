from openpyxl import load_workbook  # библиотека чтения Excel

#  Спрашивем у пользователя какой файл Excel будем открывать
from tkinter import Tk     # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename

#  Спрашивам у пользователя какой файл открыть
# we don't want a full GUI, so keep the root window from appearing
Tk().withdraw()
# show an "Open" dialog box and return the path to the selected file
filename = askopenfilename()

#  Загружаем файл exlel по полученноум ранее имени
wb = load_workbook(str(filename))
amount_sheets = len(wb.sheetnames)  # Получаем количество вкладок в Excel
sheetNumber = 1  # Номер вкладки
amountRow = 0  # Количество строк
amountCol = 0  # Количество столбцов
row_list: list = []  # Лист столбцов из файла


def reedSheet():  # Метод чтения определенной вкладки
    print('Открываем вкладку ' + '"' + wb.sheetnames[sheetNumber]+'"')
    ws = wb[wb.sheetnames[sheetNumber]]
    # Определяем количество строк и столбцов в откытом документе
    global amountRow, amountCol
    amountRow = ws.max_row
    amountCol = ws.max_column

    #  Пробегаемся по всему файлу Excel и наполняем список
    for row in ws.iter_rows():
        row = [cell.value for cell in row]
        row_list.append(row)


# Проверяем количество вкладок в таблице
if amount_sheets == 1:  # Если всего одна вкладка, сразу же читаем ее
    print('В документе найдена всего одна вкладка')
    sheetNumber = 0
    amount_sheets = 0
    reedSheet()
else:  # Если несколько вкладок, спрашивем какую нам надо
    print('В документе найдено '+str(amount_sheets)+' вкладок')
    for i in wb.sheetnames:  # Выводим названия этих вкладок
        print(str(sheetNumber) + ' ' + str(i))
        sheetNumber += 1
    print('Выбирите номер вкладки для чтения')
    sheetNumber = int(input())-1
    reedSheet()

print(f'Вкладка содержит {str(amountCol)} колонки и {str(amountRow)} строк')
for i in range(amountCol):  # Выводим название столбцов
    print(str(i+1) + ' ' + str(row_list[0][i]))
print('Выбирите номер столбца с именами')
numberRow = int(input())

pos = 0  # Счетчик позиции строки
users = []  # Список пользователей ФамилияИО
#  Генерируем ФИО
for i in row_list:
    try:  # Если ячейка имеет правильный формат
        # Разбираем на отдельно Ф И О
        cash = row_list[pos][numberRow-1].split(' ')
        if pos > 0:  # Пропускаем первую строку так как там название столбца
            # Отделяем инициалы
            users.append(f'{cash[0]} {cash[1][0:1]}.{cash[2][0:1]}.')
        pos += 1
    except Exception:  # если в ячейки что-то другое
        print('Строка '+str(pos)+' имеет не верный формат - пропускаем!')
        # Значение все равно добавляем, иначе будет сдвиг
        users.append('Не верный!')
        pos += 1

# Записываем готовый массив пользователей обратно в Excel таблицу
# в ту же вкладку в новый столбец
# Открываем файл на выбранной ранее вкладке
ws = wb[wb.sheetnames[sheetNumber]]
# Добавлеяем в первую строку название столбца
ws.cell(row=1, column=amountCol+1).value = 'ФИО'
pos = 2  # Так как отчет с 1 и первая строка - название

for row in users:
    # Вставляем в ячейку значение из Users
    ws.cell(row=pos, column=amountCol+1).value = row
    pos += 1
try:  # Если запись в файл получлась
    wb.save(str(filename))
    print('Данные успешно записаны в файл '+str(filename))
except Exception:  # Если не получилась
    print('Не удалось записать файл! '+str(filename))
